from datetime import datetime, time

from openpyxl.styles import Font, PatternFill
import openpyxl

from functions import totalHoras, tipo_dias_descuento
import buscarPlantillaXLSX

ARCHIVO_XLSX = buscarPlantillaXLSX.rutaArchivoXLS()
ListHojasCalculo = []

COL_FECHA = 'C'
COL_HORA_ENTRADA = 'D'
COL_HORA_SALIDA = 'E'
COL_TOTAL_HORAS = 'F'
COL_DESCU_ALMUERZO = 'G'

try:
    libroExcel = openpyxl.load_workbook(ARCHIVO_XLSX)
    ListHojasCalculo = libroExcel.sheetnames
    # contador del ciclo while
    contadorHojas = 0  # valor inicial debe ser 0
    while contadorHojas < len(ListHojasCalculo):
        hojaActiva = libroExcel[ListHojasCalculo[contadorHojas]]
        contadorFila = 7  # a partir de la celda 7 se calculan las horas
        for fila in hojaActiva.iter_rows():
            celdaActivaFecha = COL_FECHA + str(contadorFila)
            celdaHoraEntrada = COL_HORA_ENTRADA + str(contadorFila)
            celdaHoraSalida = COL_HORA_SALIDA + str(contadorFila)
            celdaTotalHoras = COL_TOTAL_HORAS + str(contadorFila)
            celdaDescuAlmuerzo = COL_DESCU_ALMUERZO + str(contadorFila)
            
            valorFecha = hojaActiva[celdaActivaFecha].value
            valorHoraEntrada = hojaActiva[celdaHoraEntrada].value
            valorHoraSalida = hojaActiva[celdaHoraSalida].value
            # la Celda final
            if str(valorFecha) == "TOTALES":
                print(hojaActiva)
                break
            # Confirma si la celda tiene un tipo de formato valido y si no se coloca un fondo rojo
            # con el texto "NO DATOS"
            elif type(valorHoraEntrada) is str or type(valorHoraSalida) is str:
                hojaActiva[celdaTotalHoras].value = "NO DATOS"
                hojaActiva[celdaTotalHoras].fill = PatternFill(patternType="solid",
                                                              fgColor="CA4958")
            # si la celda actual es vacia se le asigna 0 al total horas
            elif valorHoraEntrada is None or valorHoraSalida is None:
                hojaActiva[celdaTotalHoras].value = 0
                libroExcel.save(ARCHIVO_XLSX)

            else:
                hojaActiva[celdaTotalHoras].fill = PatternFill(patternType=None)  # Quita el fondo de la celda
                hojaActiva[celdaTotalHoras].font = Font(bold=False, size=14, color='00000000', name="Arial") # el formato
                # se emplea el metodo diferrenciaHoras del modulo totalHoras
                cantidaHoras = totalHoras.diferrenciaHoras(valorHoraEntrada, valorHoraSalida)
                print("valorHoraEntrada: ", valorHoraEntrada, " valorHoraSalida: ", valorHoraSalida, " Total Horas: ",
                      cantidaHoras, " ", hojaActiva)
                # asigna el valor de las horas a la celda
                hojaActiva[celdaTotalHoras].value = cantidaHoras
                # Descuenta el timepo de almuerzo
                hojaActiva[celdaDescuAlmuerzo].value = tipo_dias_descuento.descuentoAlmuerzo(cantidaHoras, valorFecha)
                # Agrega un estilo al la celda de la fecha
                if (tipo_dias_descuento.tipoDia(valorFecha) == "FIN_SEM_DOMINGO" or
                        tipo_dias_descuento.tipoDia(valorFecha) == "FESTIVO"):
                    hojaActiva[celdaActivaFecha].fill =  PatternFill(patternType="solid",
                                                              fgColor="CA4958")
                    hojaActiva[celdaActivaFecha].font = Font(name="Arial", bold=True, color='212120')
                # guarda los cambios
                libroExcel.save(ARCHIVO_XLSX)

            contadorFila = contadorFila + 1

        contadorHojas = contadorHojas + 1

except Exception as ex:
    print("Archivo corrupto \n posible falla " + str(ex) +
          "\n Hoja: ", hojaActiva, celdaHoraEntrada, valorHoraEntrada)
